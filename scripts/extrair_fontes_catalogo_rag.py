#!/usr/bin/env python3
"""
Extrai PDF/XLSX do catálogo piloto para Markdown em ``dominio_fiscal/extraido/FONTE-xxx.md``.

Requer: ``pypdf``, ``openpyxl`` (dependências do projeto).

Uso:
  PYTHONPATH=. python scripts/extrair_fontes_catalogo_rag.py
  PYTHONPATH=. python scripts/extrair_fontes_catalogo_rag.py --somente-piloto
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]

from src.infrastructure.rag.catalogo_fontes_index import (  # noqa: E402
    caminho_extraido_markdown,
    carregar_entradas_catalogo,
)


def _extrair_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    partes: list[str] = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t.strip():
            partes.append(t.strip())
    return "\n\n".join(partes)


def _extrair_xlsx(path: Path, *, max_linhas: int = 500) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    blocos: list[str] = []
    for sheet in wb.worksheets:
        blocos.append(f"## Folha: {sheet.title}")
        for count, row in enumerate(sheet.iter_rows(values_only=True)):
            if count >= max_linhas:
                blocos.append("...(truncado)")
                break
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                blocos.append(" | ".join(cells))
    wb.close()
    return "\n".join(blocos)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrai fontes binárias do catálogo para Markdown."
    )
    parser.add_argument("--somente-piloto", action="store_true", default=True)
    args = parser.parse_args()

    total = 0
    for ent in carregar_entradas_catalogo():
        if args.somente_piloto and not ent.piloto:
            continue
        bruto = ROOT / ent.caminho
        if not bruto.is_file():
            print(f"AVISO: ausente {ent.caminho}", file=sys.stderr)
            continue
        suf = bruto.suffix.lower()
        if suf in {".md", ".txt"}:
            continue
        if suf not in {".pdf", ".xlsx", ".xls"}:
            print(f"SKIP: formato {suf} {ent.id}", file=sys.stderr)
            continue
        try:
            texto = _extrair_pdf(bruto) if suf == ".pdf" else _extrair_xlsx(bruto)
        except ImportError as exc:
            print(
                f"ERRO: dependência em falta ({exc}). Rode: pip install pypdf openpyxl",
                file=sys.stderr,
            )
            sys.exit(1)
        except Exception as exc:
            print(f"ERRO: {ent.id} {bruto}: {exc}", file=sys.stderr)
            continue
        if not texto.strip():
            print(f"AVISO: texto vazio {ent.id}", file=sys.stderr)
            continue
        dest = caminho_extraido_markdown(ent.id)
        dest.parent.mkdir(parents=True, exist_ok=True)
        cabecalho = (
            f"---\n"
            f"fonte_id: {ent.id}\n"
            f"classe: {ent.classe}\n"
            f"origem: {ent.caminho}\n"
            f"titulo: {ent.titulo}\n"
            f"---\n\n"
        )
        dest.write_text(cabecalho + texto.strip() + "\n", encoding="utf-8")
        total += 1
        print(f"OK {ent.id} -> {dest.relative_to(ROOT)}")

    print(f"Extraídas: {total} fontes")


if __name__ == "__main__":
    main()
